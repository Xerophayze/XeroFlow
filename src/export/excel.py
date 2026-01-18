"""
ExportExcel.py
=================

This module provides functionality to convert GitHub‑flavoured Markdown (GFM)
into an Excel workbook.  It draws inspiration from the companion
``ExportWord.py`` module, which performs the same job for Word documents.

The converter supports many of the same constructs as the Word version:

* **Headings**, lists, block quotes and plain paragraphs.
* **Markdown tables** are detected and written into the workbook with
  header formatting.
* **Code fences** can contain plain text or special languages.  A code fence
  labelled ``chart`` will be rendered into an image via the QuickChart API,
  while a fence labelled ``mermaid`` will be rendered into a diagram via
  mermaid.ink.  Those images are both embedded into the spreadsheet and
  saved into a ``ChartsAndDiagrams`` subdirectory alongside the workbook.
* **Horizontal rules** and page breaks.  Since Excel has no concept of a
  page break within a worksheet, a ``<pbreak>`` tag will start a new
  worksheet.

Whilst Excel does not offer the same rich text capabilities as Word, this
implementation endeavours to preserve the structure of the original Markdown
and applies sensible cell styling where possible.  For example, headings are
rendered with larger bold fonts, code blocks get a monospaced font and
shaded background, and block quotes are italicised and indented.

To save the workbook, call ``convert_markdown_to_excel`` with your Markdown
string and an optional output path.  If no output path is supplied a file
save dialog will be presented (on systems with a GUI available).  The file
will be saved as an ``.xlsx`` workbook.

Example usage::

    from src.export.excel import convert_markdown_to_excel

    markdown = '''
    # Heading Example

    This is a paragraph with **bold** and _italic_ text.

    | Name | Age | Score |
    |----|----|----|
    | Alice | 30 | 95 |
    | Bob   | 25 | 80 |

    ```chart
    {
      "width": 400,
      "height": 200,
      "config": {
        "type": "bar",
        "data": {
          "labels": ["Alice", "Bob"],
          "datasets": [{ "label": "Score", "data": [95, 80] }]
        }
      }
    }
    ```
    '''

    convert_markdown_to_excel(markdown, output_path="report.xlsx", formatting_enabled=True)
"""

import logging
import os
import re
import tempfile
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as OpenPyxlImage
from datetime import datetime

# Try to import mermaid support; if unavailable we'll fall back to code blocks.
try:
    import mermaid as md
    from mermaid.graph import Graph
    MERMAID_AVAILABLE = True
except Exception:
    logging.warning("mermaid-py package not available. Mermaid diagrams will be rendered as code blocks.")
    MERMAID_AVAILABLE = False

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s [%(levelname)s] %(message)s')

###############################################################################
# Table parsing
###############################################################################

def parse_table(lines: List[str], current_index: int) -> Tuple[Optional[Dict[str, List[List[str]]]], int]:
    """Parse a GitHub‑flavoured Markdown table.

    The table parser consumes lines beginning with ``|`` as header and
    subsequent rows.  It returns a dictionary with ``headers`` and
    ``data`` keys and the index where processing should continue.  If no
    table is detected, ``None`` is returned and the index is unchanged.

    Args:
        lines: List of lines from the original Markdown.
        current_index: Index in ``lines`` to begin parsing at.

    Returns:
        (table_data, new_index): The parsed table and the index after
        the table ends.
    """
    if current_index >= len(lines):
        return None, current_index

    header_line = lines[current_index].strip()
    if not header_line.startswith('|'):
        return None, current_index

    headers = [cell.strip() for cell in header_line.strip('|').split('|')]
    current_index += 1

    # Optional separator line (---|---). Skip if present.
    if current_index < len(lines):
        separator_line = lines[current_index].strip()
        if separator_line.startswith('|') and all('-' in cell for cell in separator_line.strip('|').split('|')):
            current_index += 1

    data: List[List[str]] = []
    while current_index < len(lines):
        line = lines[current_index].strip()
        if not line.startswith('|'):
            break
        row = [cell.strip() for cell in line.strip('|').split('|')]
        if len(row) == len(headers):
            data.append(row)
        current_index += 1

    if not data:
        return None, current_index

    return {'headers': headers, 'data': data}, current_index

###############################################################################
# Chart and diagram rendering
###############################################################################

def process_chart_code(code_text: str) -> Optional[bytes]:
    """Render a chart specified in JSON using the QuickChart API.

    This implementation avoids importing the ``quickchart`` Python package,
    which may not be installed.  Instead it sends the chart definition
    directly to the public QuickChart API endpoint.  Prior to submission
    the JSON is sanitised to remove callback functions and other
    unsupported constructs.

    Args:
        code_text: A JSON string describing the chart configuration.  The
            top‑level keys may include ``width`` and ``height`` along with
            ``config`` containing the actual Chart.js configuration.

    Returns:
        The image as bytes, or ``None`` if the API call fails.
    """
    import json
    try:
        logging.info("[DEBUG Chart] Processing chart code")

        # Remove JavaScript callback functions to make the JSON safe for parsing
        code_text = re.sub(r'"callback":\s*"function\(.*?\)\s*{[^}]*}"', '', code_text, flags=re.DOTALL)
        code_text = re.sub(r'"formatter":\s*\(.*?\)\s*=>\s*{[^}]*},?', '', code_text, flags=re.DOTALL)
        code_text = re.sub(r'"tooltips":\s*{[^}]*"callbacks":\s*{[^}]*}}', '', code_text, flags=re.DOTALL)
        code_text = re.sub(r'"annotation":\s*{[^}]*"annotations":\s*\[[^]]*\]}', '', code_text, flags=re.DOTALL)
        code_text = re.sub(r',\s*}', '}', code_text)

        config = json.loads(code_text)
        width = config.get('width', 500)
        height = config.get('height', 300)
        chart_config = config.get('config', {})
        post_data = {
            'chart': chart_config,
            'width': width,
            'height': height,
            'format': 'png'
        }
        headers = {'Content-Type': 'application/json', 'Accept': 'image/png'}
        response = requests.post('https://quickchart.io/chart', json=post_data, headers=headers, timeout=30)
        if response.headers.get('Content-Type', '').startswith('image/'):
            return response.content
        logging.error(f"Unexpected response from QuickChart: {response.status_code}")
        return None
    except Exception as e:
        logging.error(f"Error processing chart code: {e}")
        return None

def process_mermaid_diagram(mermaid_code: str) -> Optional[bytes]:
    """Render a Mermaid diagram into a high-resolution image using Kroki SVG.

    Uses Kroki's SVG endpoint (vector-based, no width limits) and converts to
    high-resolution PNG at 2400px width. Falls back to mermaid.ink if Kroki fails.

    Args:
        mermaid_code: The Mermaid diagram definition.

    Returns:
        The image as bytes (JPEG format), or ``None`` if rendering fails.
    """
    if not MERMAID_AVAILABLE:
        logging.warning("Mermaid package not available. Cannot render diagram.")
        return None
    
    # Try Kroki SVG first (no width limitations)
    try:
        kroki_base = os.environ.get("KROKI_SERVER", "https://kroki.io").rstrip('/')
        kroki_svg_url = f"{kroki_base}/mermaid/svg"
        headers = {
            'Accept': 'image/svg+xml',
            'Content-Type': 'text/plain'
        }
        logging.info(f"Attempting Kroki SVG rendering for Excel...")
        response = requests.post(kroki_svg_url, data=mermaid_code.encode('utf-8'), headers=headers, timeout=30)
        
        if response.status_code == 200 and response.headers.get('Content-Type', '').startswith('image/svg'):
            logging.info(f"Successfully generated Mermaid SVG via Kroki")
            
            # Convert SVG to high-resolution PNG using Kroki's PNG endpoint
            # This avoids SVG compatibility issues
            try:
                kroki_png_url = f"{kroki_base}/mermaid/png"
                png_headers = {
                    'Accept': 'image/png',
                    'Content-Type': 'text/plain'
                }
                logging.info("Converting to high-resolution PNG via Kroki...")
                png_response = requests.post(kroki_png_url, data=mermaid_code.encode('utf-8'), headers=png_headers, timeout=30)
                
                if png_response.status_code == 200 and png_response.headers.get('Content-Type', '').startswith('image/'):
                    from PIL import Image
                    
                    png_image = Image.open(BytesIO(png_response.content))
                    logging.info(f"Kroki PNG size: {png_image.size[0]}x{png_image.size[1]} pixels")
                    
                    # Convert to RGB for JPEG
                    if png_image.mode in ('RGBA', 'LA', 'P'):
                        rgb_image = Image.new('RGB', png_image.size, (255, 255, 255))
                        if png_image.mode == 'P':
                            png_image = png_image.convert('RGBA')
                        rgb_image.paste(png_image, mask=png_image.split()[-1] if png_image.mode in ('RGBA', 'LA') else None)
                        png_image = rgb_image
                    elif png_image.mode != 'RGB':
                        png_image = png_image.convert('RGB')
                    
                    # Save as JPEG with 80% quality
                    buf = BytesIO()
                    png_image.save(buf, format='JPEG', quality=80, optimize=True, progressive=True)
                    jpg_bytes = buf.getvalue()
                    
                    logging.info(f"Converted to JPEG: {len(jpg_bytes):,} bytes (quality=80)")
                    return jpg_bytes
                else:
                    logging.warning(f"Kroki PNG conversion failed: {png_response.status_code}")
            except Exception as e:
                logging.error(f"Error converting via Kroki PNG: {e}")
        else:
            logging.warning(f"Kroki SVG failed. Status: {response.status_code}. Falling back to mermaid.ink...")
    except Exception as e:
        logging.error(f"Error using Kroki: {e}. Falling back to mermaid.ink...")
    
    # Fallback to mermaid.ink (limited to ~784px width)
    try:
        import base64
        graphbytes = mermaid_code.encode('utf8')
        base64_string = base64.urlsafe_b64encode(graphbytes).decode('ascii')
        url = f"https://mermaid.ink/img/{base64_string}"
        
        logging.info("Attempting mermaid.ink fallback...")
        response = requests.get(url, timeout=15)
        if response.status_code == 200 and response.headers.get('Content-Type', '').startswith('image/'):
            logging.info(f"Successfully generated via mermaid.ink (width limited to ~784px)")
            try:
                from PIL import Image
                img = Image.open(BytesIO(response.content))
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                buf = BytesIO()
                img.save(buf, format='JPEG', quality=80, optimize=True)
                jpeg = buf.getvalue()
                if len(jpeg) < len(response.content):
                    return jpeg
                return response.content
            except Exception:
                return response.content
        logging.error(f"mermaid.ink also failed. Status: {response.status_code}")
        return None
    except Exception as e:
        logging.error(f"Error processing Mermaid diagram: {e}")
        return None

def save_image_to_subdirectory(image_bytes: bytes, output_path: str, image_type: str, image_index: int) -> Optional[str]:
    """Save chart or diagram images next to the workbook.

    Images generated from charts and mermaid diagrams are saved into a
    ``ChartsAndDiagrams`` folder alongside the Excel file.  The filename
    includes the workbook name, image type and an index.

    Args:
        image_bytes: The raw image bytes.
        output_path: The path where the workbook will be saved.
        image_type: Either ``'chart'`` or ``'mermaid'``.
        image_index: A running counter used to create unique filenames.

    Returns:
        The absolute path to the saved image, or ``None`` on failure.
    """
    try:
        if not image_bytes or not output_path:
            return None
        doc_dir = os.path.dirname(os.path.abspath(output_path))
        charts_dir = os.path.join(doc_dir, 'ChartsAndDiagrams')
        os.makedirs(charts_dir, exist_ok=True)
        doc_name = os.path.splitext(os.path.basename(output_path))[0]
        
        # Detect file type from image bytes
        if image_bytes.startswith(b'<svg') or image_bytes.startswith(b'<?xml'):
            ext = 'svg'
        elif image_bytes.startswith(b'\x89PNG'):
            ext = 'png'
        elif image_bytes.startswith(b'\xff\xd8\xff'):
            ext = 'jpg'
        else:
            # Default based on image type
            ext = 'svg' if image_type == 'mermaid' else 'png'
            
        image_filename = f"{doc_name}_{image_type}_{image_index:03d}.{ext}"
        image_path = os.path.join(charts_dir, image_filename)
        with open(image_path, 'wb') as f:
            f.write(image_bytes)
        logging.info(f"Saved {image_type} image to: {image_path}")
        return image_path
    except Exception as e:
        logging.error(f"Failed to save {image_type} image: {e}")
        return None

###############################################################################
# Main conversion function
###############################################################################

def convert_markdown_to_excel(markdown_text: str, output_path: Optional[str] = None, *, formatting_enabled: bool = False) -> None:
    """Convert GitHub‑flavoured Markdown into an Excel workbook.

    The function parses the supplied Markdown line by line and writes
    corresponding entries into a workbook.  When ``formatting_enabled`` is
    ``True`` the converter will recognise headings, lists, code blocks,
    tables and more, applying simple Excel formatting where appropriate.
    Charts and Mermaid diagrams inside code fences are rendered via
    external services and inserted into the sheet.

    Args:
        markdown_text: The Markdown content to convert.
        output_path: Where to save the workbook.  If omitted a file save
            dialog will be presented (if available).
        formatting_enabled: If ``True`` markup will be interpreted.  When
            ``False`` the raw Markdown will be written into a single cell.
    """
    # Lazy import of Tkinter so that headless environments can still run
    # convert_markdown_to_excel without requiring a display.
    try:
        from tkinter import filedialog, Tk, messagebox  # type: ignore
    except Exception:
        filedialog = None
        Tk = None
        messagebox = None

    logging.info("[DEBUG ExportExcel] Starting conversion")
    logging.info(f"Formatting enabled: {formatting_enabled}")
    logging.info(f"Output path: {output_path}")

    # Determine the save path
    file_path = output_path
    if file_path is None:
        if filedialog is None:
            raise RuntimeError("No output_path specified and Tkinter file dialog not available.")
        root = Tk()
        root.withdraw()
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")], title="Save as")
        if not file_path:
            if messagebox:
                messagebox.showwarning("No Save Location", "No file was selected. Operation cancelled.")
            return

    try:
        wb = Workbook()
        sheet_index = 0
        ws = wb.active
        ws.title = f"Sheet{sheet_index + 1}"
        current_row = 1

        # Counters for numbering lists and for images
        list_counters: Dict[int, int] = {}
        chart_counter = 0
        mermaid_counter = 0

        if not formatting_enabled:
            # Write the raw Markdown into a single cell
            ws.cell(row=current_row, column=1, value=markdown_text)
        else:
            lines = markdown_text.split('\n')
            index = 0
            in_code_block = False
            code_block_language = ''
            code_block_content: List[str] = []
            while index < len(lines):
                line = lines[index]
                stripped = line.strip()

                # Handle end of code fence
                if in_code_block:
                    if stripped.startswith('```'):
                        code_text = '\n'.join(code_block_content)
                        if code_block_language == 'chart':
                            image_bytes = process_chart_code(code_text)
                            if image_bytes:
                                chart_counter += 1
                                image_path = save_image_to_subdirectory(image_bytes, file_path, 'chart', chart_counter)
                                # Insert image into worksheet
                                if image_path:
                                    img = OpenPyxlImage(image_path)
                                    ws.add_image(img, f"A{current_row}")
                                    # Adjust row height to accommodate image (approximate)
                                    ws.row_dimensions[current_row].height = img.height * 0.75 / 1.5
                                    # Reserve a few rows after the image
                                    current_row += int(img.height / 15) + 1
                                else:
                                    # Fallback: write the JSON into a grey box
                                    _write_code_block(ws, code_text, current_row)
                                    current_row += 1
                            else:
                                _write_code_block(ws, code_text, current_row)
                                current_row += 1
                        elif code_block_language == 'mermaid':
                            image_bytes = process_mermaid_diagram(code_text)
                            if image_bytes:
                                mermaid_counter += 1
                                image_path = save_image_to_subdirectory(image_bytes, file_path, 'mermaid', mermaid_counter)
                                if image_path:
                                    img = OpenPyxlImage(image_path)
                                    ws.add_image(img, f"A{current_row}")
                                    ws.row_dimensions[current_row].height = img.height * 0.75 / 1.5
                                    current_row += int(img.height / 15) + 1
                                else:
                                    _write_code_block(ws, code_text, current_row)
                                    current_row += 1
                            else:
                                _write_code_block(ws, code_text, current_row)
                                current_row += 1
                        else:
                            _write_code_block(ws, code_text, current_row)
                            current_row += 1
                        in_code_block = False
                        code_block_language = ''
                        code_block_content.clear()
                        index += 1
                        continue
                    else:
                        code_block_content.append(line)
                        index += 1
                        continue

                # Start of a code fence
                if stripped.startswith('```'):
                    in_code_block = True
                    code_block_language = stripped[3:].strip().lower()
                    code_block_content.clear()
                    index += 1
                    continue

                # Page break
                if stripped == '<pbreak>':
                    sheet_index += 1
                    ws = wb.create_sheet(title=f"Sheet{sheet_index + 1}")
                    current_row = 1
                    list_counters.clear()
                    index += 1
                    continue

                # Horizontal rule
                if re.match(r'^(\*\*\*|---|___)$', stripped):
                    # Represent a horizontal rule as an empty row with a bottom border
                    cell = ws.cell(row=current_row, column=1, value='')
                    cell.border = cell.border.copy(bottom=cell.border.bottom.copy(style='thin'))
                    current_row += 1
                    index += 1
                    continue

                # Parse table
                if stripped.startswith('|') and stripped.endswith('|'):
                    parsed, new_index = parse_table(lines, index)
                    if parsed:
                        _write_table(ws, parsed, current_row)
                        current_row += len(parsed['data']) + 1  # header + data
                        index = new_index
                        continue

                # Block quotes
                if stripped.startswith('>'):
                    quote_match = re.match(r'^(>+)(.*)$', stripped)
                    if quote_match:
                        level = len(quote_match.group(1))
                        content = quote_match.group(2).strip()
                        cell = ws.cell(row=current_row, column=1 + level - 1, value=content)
                        cell.font = Font(italic=True, color='555555')
                        cell.alignment = Alignment(wrap_text=True)
                        current_row += 1
                        index += 1
                        continue

                # Blank line
                if not stripped:
                    current_row += 1
                    list_counters.clear()
                    index += 1
                    continue

                # Headings
                heading_match = re.match(r'^(#{1,6})\s+(.*)', stripped)
                if heading_match:
                    level = len(heading_match.group(1))
                    content = heading_match.group(2).strip()
                    # Write heading into column 1 with styling
                    cell = ws.cell(row=current_row, column=1, value=content)
                    size_map = {1: 20, 2: 18, 3: 16, 4: 14, 5: 12, 6: 11}
                    cell.font = Font(bold=True, size=size_map.get(level, 10), color='002366')
                    current_row += 1
                    list_counters.clear()
                    index += 1
                    continue

                # Lists (ordered and unordered)
                list_match = re.match(r'^(\s*)([-*+]|(\d+\.))\s+(.*)', line)
                if list_match:
                    indent_str, bullet, content = list_match.group(1), list_match.group(2), list_match.group(4)
                    level = len(indent_str) // 2
                    if bullet.endswith('.') and bullet[:-1].isdigit():
                        # Ordered list
                        lvl = level
                        if lvl not in list_counters:
                            list_counters[lvl] = 1
                        number = list_counters[lvl]
                        list_counters[lvl] += 1
                        # Reset deeper levels
                        for k in list(list_counters.keys()):
                            if k > lvl:
                                list_counters.pop(k)
                        bullet_text = f"{number}. "
                    else:
                        bullet_text = '• '
                    cell = ws.cell(row=current_row, column=1 + level, value=bullet_text + content)
                    cell.alignment = Alignment(wrap_text=True)
                    current_row += 1
                    index += 1
                    continue

                # Regular paragraph: attempt to infer data type for single-cell paragraphs
                value, fmt = _infer_cell_value_and_format(stripped)
                cell = ws.cell(row=current_row, column=1, value=value)
                if fmt:
                    cell.number_format = fmt
                cell.alignment = Alignment(wrap_text=True)
                current_row += 1
                list_counters.clear()
                index += 1

            # If code fence never closed
            if in_code_block and code_block_content:
                code_text = '\n'.join(code_block_content)
                _write_code_block(ws, code_text, current_row)
                current_row += 1

        # Auto‑adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = str(cell.value) if cell.value is not None else ''
                if len(value) > max_length:
                    max_length = len(value)
            ws.column_dimensions[col_letter].width = max(10, min(max_length + 2, 60))

        wb.save(file_path)
        logging.info(f"[DEBUG ExportExcel] Workbook saved to {file_path}")
    except Exception as e:
        logging.error(f"An error occurred during conversion: {e}")
        if messagebox:
            messagebox.showerror("Error", f"An error occurred: {e}")

###############################################################################
# Helper functions
###############################################################################

def _write_table(ws, table_data: Dict[str, List[List[str]]], start_row: int) -> None:
    """Write a Markdown table to the worksheet starting at ``start_row``.

    Headers are written in bold.  Data rows follow immediately below.  Each
    cell value is passed through a type inference routine that attempts to
    convert strings into appropriate Python objects (numbers, dates,
    percentages, currency values or formulas).  When a conversion is
    successful a suitable number format is also applied to the cell.

    Args:
        ws: The worksheet to write to.
        table_data: A dictionary with ``headers`` and ``data`` keys.
        start_row: The row index in the worksheet to begin writing at.
    """
    headers = table_data.get('headers', [])
    data = table_data.get('data', [])
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    # Identify columns that should be formatted as percentages based on header text
    percent_columns = set()
    for col_index, header in enumerate(headers, start=1):
        hl = str(header).lower() if header is not None else ''
        if ('%' in str(header)) or any(tok in hl for tok in ('percent', 'rate', 'ratio', 'margin')):
            percent_columns.add(col_index)
    # Identify columns that should be formatted as currency based on header text
    currency_columns = set()
    currency_header_tokens = (
        'amount', 'price', 'cost', 'budget', 'actual', 'income', 'expense',
        'revenue', 'total', 'balance', 'sales', 'profit', 'fee', 'charge'
    )
    for col_index, header in enumerate(headers, start=1):
        hl = str(header).lower() if header is not None else ''
        if any(tok in hl for tok in currency_header_tokens) or any(sym in str(header) for sym in ('$','€','£','¥')):
            currency_columns.add(col_index)

    # Identify columns to force as text (avoid numeric auto-conversion) based on header
    text_columns = set()
    text_header_tokens = (
        'id', 'code', 'sku', 'account', 'name', 'description', 'notes', 'category', 'type'
    )
    for col_index, header in enumerate(headers, start=1):
        hl = str(header).lower() if header is not None else ''
        if any(tok in hl for tok in text_header_tokens):
            text_columns.add(col_index)

    # Offset to translate A1-style formulas authored for a table starting at row 1
    # (header at row 1, first data row at row 2) into the actual worksheet rows.
    # Example: if start_row == 8, then B2 -> B(2 + 7) -> B9
    base_offset = start_row - 1
    for row_offset, row_data in enumerate(data, start=1):
        current_sheet_row = start_row + row_offset
        # Heuristic: treat rows with 'total' or 'summary' text as aggregate rows
        row_texts = [str(x).lower() for x in row_data if isinstance(x, str)]
        is_aggregate_row = any(('total' in t) or ('summary' in t) for t in row_texts)
        for col_index, raw_value in enumerate(row_data, start=1):
            # Check for inline cell-wide markdown styling first
            styled_text = None
            is_bold = False
            is_italic = False
            if isinstance(raw_value, str):
                styled_text, is_bold, is_italic = _parse_inline_markdown_style(raw_value)

            value, number_format = _infer_cell_value_and_format(raw_value, row_offset=base_offset)

            # If the raw cell used markdown to style the entire content, keep it as text
            # so the style can be applied (do this before formula handling)
            if styled_text is not None and (is_bold or is_italic):
                value = styled_text
                number_format = None

            # Smart rebase: for simple same-row formulas, align row refs to the current row.
            # This is now attempted for ALL rows. The helper safely ignores range-based totals
            # (contains ':') and formulas with absolute rows ('$'), so aggregate rows remain intact
            # unless the formula is a simple same-row expression (e.g., E5/C5), which we do want
            # to correct to the proper row.
            if isinstance(value, str) and value.startswith('='):
                value = _smart_rebase_row_formula(value, current_sheet_row)
            cell = ws.cell(row=current_sheet_row, column=col_index, value=value)
            if number_format:
                cell.number_format = number_format
            # Apply percent formatting for designated columns regardless of prior numeric format
            if col_index in percent_columns:
                cell.number_format = '0.00%'
            # Apply currency formatting when header suggests a currency column (even for formulas)
            if col_index in currency_columns:
                existing_nf = cell.number_format or ''
                has_currency_symbol = any(sym in str(existing_nf) for sym in ('$','€','£','¥'))
                if '%' not in existing_nf:
                    if not existing_nf or (('#' in existing_nf or '0' in existing_nf) and not has_currency_symbol):
                        cell.number_format = '$#,##0.00'
            # Force text for selected columns (except formulas)
            if col_index in text_columns and not (isinstance(value, str) and value.startswith('=')):
                # Ensure the cell is stored as text
                cell.value = '' if value is None else str(value)
                cell.number_format = '@'

            # Apply inline styles if present
            if styled_text is not None and (is_bold or is_italic):
                cell.font = Font(bold=is_bold, italic=is_italic)

            # Alignment: right-align numbers/currency/percent/formulas, left-align text
            align_right = False
            if isinstance(cell.value, (int, float)):
                align_right = True
            if isinstance(cell.value, str) and cell.value.startswith('='):
                align_right = True
            nf = getattr(cell, 'number_format', None)
            if nf and (('%' in nf) or ('0' in nf) or ('#' in nf)):
                align_right = True
            cell.alignment = Alignment(horizontal=('right' if align_right else 'left'), wrap_text=not align_right)

def _write_code_block(ws, code_text: str, start_row: int) -> None:
    """Insert a code block into the worksheet with monospaced font and shading.

    The code block is written into a single cell spanning multiple lines.
    A grey background and Consolas font are applied.  The cell is wrapped
    automatically.

    Args:
        ws: The worksheet to write to.
        code_text: The text of the code block.
        start_row: The row index at which to insert the block.
    """
    cell = ws.cell(row=start_row, column=1, value=code_text)
    cell.font = Font(name='Consolas', size=10)
    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    cell.alignment = Alignment(wrap_text=True)


# -----------------------------------------------------------------------------
# Type inference for table cells
# -----------------------------------------------------------------------------

def _adjust_formula_row_refs(formula: str, row_offset: int) -> str:
    """Adjust A1-style row numbers in a formula by ``row_offset``.

    Only the numeric row part of cell references is shifted. Column letters and
    any absolute markers (e.g., ``$B$2``) are preserved; absolute rows are not
    shifted. Ranges like ``G2:G4`` are handled because both ends match.
    """
    if not formula or row_offset == 0:
        return formula
    # Match cell references like B2, $B2, B$2, $B$2, including up to 3 letters
    cell_ref_re = re.compile(r"(\$?[A-Za-z]{1,3})(\$?)(\d+)")

    def repl(m: re.Match) -> str:
        col = m.group(1)  # may include leading $
        abs_row = m.group(2)  # '$' if absolute row
        row_num = int(m.group(3))
        if abs_row == '$':
            # Respect absolute row references
            return f"{col}{abs_row}{row_num}"
        return f"{col}{row_num + row_offset}"

    return cell_ref_re.sub(repl, formula)


def _smart_rebase_row_formula(formula: str, target_row: int) -> str:
    """Heuristically rebase simple same-row formulas to the current sheet row.

    This addresses cases where the source uses hard-coded row numbers (e.g., D18-C18 or
    IF(C18=0,0,D18/C18)) that became misaligned due to inserted section rows.

    Rules:
    - Only touch simple same-row patterns (Dn-Cn, Dn+Cn, Dn/Cn, Dn*Cn) and IF(Cn=0,0,Dn/Cn).
    - Skip if formula appears to contain ranges (":") or absolute rows ("$<row>").
    - Safe to run for all rows, including aggregate/total rows; range-based totals remain unchanged.
    """
    if not formula or ':' in formula:
        return formula
    # If there are absolute row markers, leave as-is
    if re.search(r"\$\d+", formula):
        return formula
    f = formula.strip()
    # Pattern: =Dn op Cn  (op in - + / *) with same row n on both sides
    bin_op = re.compile(r"^=\s*([A-Za-z]{1,3})(\d+)\s*([\-\+\*/])\s*([A-Za-z]{1,3})(\d+)\s*$", re.IGNORECASE)
    m = bin_op.match(f)
    if m and m.group(2) == m.group(5):
        col1, row1, op, col2, _ = m.groups()
        return f"={col1}{target_row}{op}{col2}{target_row}"
    # Pattern: =IF(Cn=0,0,Dn/Cn) allowing spaces
    if_pat = re.compile(r"^=\s*IF\(\s*([A-Za-z]{1,3})(\d+)\s*=\s*0\s*,\s*0\s*,\s*([A-Za-z]{1,3})(\d+)\s*/\s*([A-Za-z]{1,3})(\d+)\s*\)\s*$", re.IGNORECASE)
    m2 = if_pat.match(f)
    if m2:
        c1, r1, c2, r2, c3, r3 = m2.groups()
        # Only rebase if all referenced rows are the same (typical same-row formula)
        if r1 == r2 == r3:
            return f"=IF({c1}{target_row}=0,0,{c2}{target_row}/{c3}{target_row})"
    return formula


def _parse_inline_markdown_style(s: str) -> Tuple[Optional[str], bool, bool]:
    """Parse simple whole-cell Markdown styles and return (text, bold, italic).

    Recognises only cases where the entire cell content is wrapped, e.g.:
    - ***text*** or ___text___  -> bold and italic
    - **text** or __text__      -> bold
    - *text* or _text_          -> italic

    Returns (None, False, False) if no whole-cell style is detected.
    """
    if s is None:
        return None, False, False
    t = str(s).strip()
    # bold+italic: ***text*** or ___text___
    m = re.match(r"^(?:\*\*\*|___)(.+?)(?:\*\*\*|___)$", t)
    if m:
        return m.group(1).strip(), True, True
    # bold: **text** or __text__
    m = re.match(r"^(?:\*\*|__)(.+?)(?:\*\*|__)$", t)
    if m:
        return m.group(1).strip(), True, False
    # italic: *text* or _text_
    m = re.match(r"^(?:\*|_)(.+?)(?:\*|_)$", t)
    if m:
        return m.group(1).strip(), False, True
    return None, False, False


def _infer_cell_value_and_format(raw_value: str, row_offset: int = 0) -> Tuple[object, Optional[str]]:
    """Attempt to infer the Python value and number format for a cell.

    Given a raw string from a Markdown table cell, this function tries to
    convert it into a more appropriate Python type.  It supports formulas
    (strings starting with ``=``), booleans, dates in a few common formats,
    percentages, currency values (USD/EUR/GBP/JPY) and plain numbers with
    optional thousand separators.  When a conversion is performed a suitable
    ``number_format`` string is returned; otherwise the format is ``None``.

    Args:
        raw_value: The cell content as a string.

    Returns:
        A tuple of ``(value, number_format)`` where ``value`` is the object to
        assign to the cell and ``number_format`` is an optional format string
        to apply to the cell.  If ``value`` is a formula (string starting
        with '=') then no format is returned.
    """
    if raw_value is None:
        return '', None
    s = str(raw_value).strip()
    # Formula detection
    if s.startswith('='):
        # Adjust A1-style row numbers by row_offset (used for tables)
        adjusted = _adjust_formula_row_refs(s, row_offset)
        return adjusted, None
    # Boolean detection
    lower = s.lower()
    if lower in ('true', 'false', 'yes', 'no'):
        return (lower in ('true', 'yes')), None
    # Date detection
    date_formats = ['%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%d/%m/%Y', '%d/%m/%y', '%Y/%m/%d']
    for fmt in date_formats:
        try:
            dt = datetime.strptime(s, fmt).date()
            # Format as ISO date
            return dt, 'yyyy-mm-dd'
        except ValueError:
            pass
    # Percentage detection
    if s.endswith('%') and len(s) > 1:
        num_str = s[:-1].strip()
        try:
            # Remove thousands separators
            num_clean = num_str.replace(',', '')
            value = float(num_clean) / 100
            decimals = 0
            if '.' in num_str:
                decimals = len(num_str.split('.')[1])
            if decimals > 0:
                fmt = '0.' + ('0' * decimals) + '%'
            else:
                fmt = '0%'
            return value, fmt
        except ValueError:
            pass
    # Currency detection
    currency_symbols = ['$', '€', '£', '¥']
    symbol = None
    stripped_number = s
    for sym in currency_symbols:
        if s.startswith(sym):
            symbol = sym
            stripped_number = s[len(sym):].strip()
            break
        elif s.endswith(sym):
            symbol = sym
            stripped_number = s[:-len(sym)].strip()
            break
    if symbol:
        try:
            # Remove thousands separators
            num_clean = stripped_number.replace(',', '')
            value = float(num_clean)
            # Determine decimals
            decimals = 0
            if '.' in stripped_number:
                decimals = len(stripped_number.split('.')[-1])
            # Create a currency format string, e.g. "$#,##0.00" or "€#,##0.0"
            # Excel uses locale settings for currency; specifying explicit symbol works.
            if decimals > 0:
                fmt = f'{symbol}#,##0.' + ('0' * decimals)
            else:
                fmt = f'{symbol}#,##0'
            return value, fmt
        except ValueError:
            pass
    # Pure numeric detection (integers or floats)
    # Remove thousands separators for parsing
    try:
        num_clean = s.replace(',', '')
        # Check if there is a decimal point
        if '.' in num_clean:
            value = float(num_clean)
            decimals = len(num_clean.split('.')[-1])
            fmt = '#,##0.' + ('0' * decimals)
            return value, fmt
        else:
            value = int(num_clean)
            fmt = '#,##0'
            return value, fmt
    except ValueError:
        pass
    # Fallback: return original string
    return s, None
