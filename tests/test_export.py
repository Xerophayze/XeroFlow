from pathlib import Path
from src.export.excel import convert_markdown_to_excel
from openpyxl import load_workbook

SRC = Path("advanced_excel_markdown_examples (1).txt")
OUTDIR = Path("test_outputs")
OUTDIR.mkdir(exist_ok=True)
OUT = OUTDIR / "advanced_examples.xlsx"

# Export
md = SRC.read_text(encoding="utf-8")
convert_markdown_to_excel(md, str(OUT), formatting_enabled=True)
print(f"Wrote: {OUT}")

# Verify formulas in appended budget example
wb = load_workbook(str(OUT))
ws = wb.active

# Find the appended budget table by header signature
headers = [
    "Category",
    "Type",
    "Budget ($)",
    "Actual ($)",
    "Variance ($)",
    "Variance %",
    "Notes",
]

header_row_idx = None
header_col_map = {}
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    values = [c.value if c.value is not None else "" for c in row]
    if values[:len(headers)] == headers:
        header_row_idx = row[0].row
        header_col_map = {name: idx+1 for idx, name in enumerate(headers)}
        break

if header_row_idx is None:
    print("Budget headers not found")
    raise SystemExit(1)

col_var_pct = header_col_map["Variance %"]

# Rows we want to check in the appended table
targets = [
    "Total Income",
    "Total Fixed Exp.",
    "Total Var. Exp.",
    "Total Savings/Debt",
    "Total Budgeted Exp.",
    "Net Cash Flow",
]

found = {}
for row in ws.iter_rows(min_row=header_row_idx+1, max_row=ws.max_row):
    name = row[0].value
    if isinstance(name, str):
        n = name.strip().strip(":")
        if n in targets and n not in found:
            r = row[0].row
            cell = ws.cell(row=r, column=col_var_pct)
            found[n] = (r, cell.value)
            if len(found) == len(targets):
                break

for k in targets:
    r, f = found.get(k, (None, None))
    print(f"{k} -> row {r}, Variance % formula: {f}")
